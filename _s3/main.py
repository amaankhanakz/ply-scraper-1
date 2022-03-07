'''
usage:
-> fill everything in env.ref.py and rename it to env.py
-> run this script
'''

import env
import boto3
import requests
import os
from shutil import rmtree
from uuid import uuid4
from openpyxl import load_workbook

def upload_img(path, url_col):
    for item in os.listdir(path):
        item_path = os.path.join(path, item)
        if os.path.isfile(item_path):
            workbook = load_workbook(filename=item_path)
            sheet = workbook.active
            for i in range(2, sheet.max_row + 1):
                sheet.cell(row=i, column=url_col).value = push_img(sheet.cell(row=i, column=url_col).value)
                workbook.save(item_path)
        if os.path.isdir(item_path):
            upload_img(item_path, url_col)

def push_img(url):
    download_location = env.download_location
    img_name = f'{uuid4()}.{env.img_type}'
    img_path = os.path.join(download_location, img_name)

    if os.path.exists(download_location) == False:
        os.makedirs(download_location)

    try:
        if img_path.endswith(env.img_type):
            with open(img_path, 'wb') as file:
                img = requests.get(url)
                file.write(img.content)
                print(f'> ⬇️  {url}')
        else:
            raise Exception('invalid link')
    except Exception as e:
        return 'unavailable'

    aws_filename = os.path.join(env.aws_file_path, img_name)
    public_url = upload_aws(img_path, aws_filename, env.aws_access_key, env.aws_secret_key, env.aws_bucket)

    return public_url

def upload_aws(local_filename, aws_filename, access_key, secret_key, bucket_name):
    s3 = boto3.client(
        's3', 
        region_name=env.aws_region_name, 
        aws_access_key_id=access_key, 
        aws_secret_access_key=secret_key
    )

    s3.upload_file(local_filename, bucket_name, aws_filename, ExtraArgs={'ContentType': f'image/{env.img_type}'})

    print(f'> ✅ {local_filename} --> [{bucket_name}] {aws_filename}')

    public_url = s3.generate_presigned_url(
                'get_object', 
                Params = {
                    'Bucket': bucket_name, 
                    'Key': aws_filename
                    }
                )

    return public_url


if __name__ == '__main__':
    upload_img(env.sheet_path, env.url_col)
    rmtree(env.download_location)
